from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.models import Submission, SubmissionRow, SubmissionSelection, TorobMatch
from app.services.excel import build_export_xlsx, parse_products_xlsx


def make_xlsx(headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_parse_standard_template_headers() -> None:
    content = make_xlsx(
        ["product_name", "price", "barcode", "brand", "description"],
        [["رب گوجه روژین", "165,000", "123", "روژین", "قوطی ۸۰۰ گرم"]],
    )

    rows = parse_products_xlsx(content)

    assert len(rows) == 1
    assert rows[0].product_name == "رب گوجه روژین"
    assert rows[0].price == "165000"
    assert rows[0].barcode == "123"
    assert rows[0].brand == "روژین"
    assert rows[0].error_message is None


def test_parse_farsi_headers_and_invalid_empty_product_name() -> None:
    content = make_xlsx(
        ["نام کالا", "قیمت", "بارکد", "برند"],
        [[None, "100000", "123", "برند تست"]],
    )

    rows = parse_products_xlsx(content)

    assert len(rows) == 1
    assert rows[0].error_message == "نام محصول خالی است."


def test_parse_farsi_headers_after_intro_row() -> None:
    content = make_xlsx(
        ["گزارش محصولات"],
        [["نام كالا", "قيمت"], ["رب گوجه روژین", "165000"]],
    )

    rows = parse_products_xlsx(content)

    assert len(rows) == 1
    assert rows[0].input_row == 3
    assert rows[0].product_name == "رب گوجه روژین"
    assert rows[0].price == "165000"


def test_parse_fee_sale_header_as_price() -> None:
    content = make_xlsx(
        ["نام محصول", "فی فروش"],
        [["رب گوجه روژین", "165000"]],
    )

    rows = parse_products_xlsx(content)

    assert rows[0].product_name == "رب گوجه روژین"
    assert rows[0].price == "165000"


def test_parse_missing_name_column_returns_error_row() -> None:
    content = make_xlsx(["قیمت"], [["100000"]])

    rows = parse_products_xlsx(content)

    assert rows[0].input_row == 1
    assert "ستون نام محصول" in rows[0].error_message


def test_parse_limits_to_500_rows() -> None:
    content = make_xlsx(["product_name"], [[f"item {index}"] for index in range(501)])

    rows = parse_products_xlsx(content)

    assert len([row for row in rows if row.product_name]) == 500
    assert "فقط 500 ردیف" in rows[-1].error_message


def test_build_export_xlsx_contains_standard_columns() -> None:
    submission = Submission(id=7, store_name="فروشگاه تست", shop_id="411147")
    row = SubmissionRow(input_row=2, input_product_name="رب", input_price="100000")
    match = TorobMatch(
        base_prk="abc",
        name="رب ترب",
        price=110000,
        price_text="از ۱۱۰۰۰۰",
        image_url="https://image.example/a.jpg",
        product_url="https://torob.com/p/abc",
        rank=0,
    )
    row.selections = [SubmissionSelection(match=match, final_price="120000")]

    content = build_export_xlsx(submission, [row])
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active

    assert sheet["A1"].value == "submission_id"
    assert sheet["H1"].value == "selected_source"
    assert sheet["H2"].value == "torob"
    assert sheet["I2"].value == "abc"
    assert sheet["L2"].value == "120000"


def test_build_export_xlsx_writes_multiple_selections_for_one_input_row() -> None:
    submission = Submission(id=8, store_name="فروشگاه تست", shop_id="411147")
    row = SubmissionRow(input_row=4, input_product_name="بالم لب", input_price="90000")
    first = TorobMatch(base_prk="first", name="بالم اول", price=80000, rank=0)
    second = TorobMatch(base_prk="second", name="بالم دوم", price=85000, rank=1)
    row.selections = [
        SubmissionSelection(match=first, final_price="99000"),
        SubmissionSelection(match=second, final_price="99000"),
    ]

    content = build_export_xlsx(submission, [row])
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active

    assert sheet["E2"].value == 4
    assert sheet["H2"].value == "torob"
    assert sheet["I2"].value == "first"
    assert sheet["I3"].value == "second"
    assert sheet["L2"].value == "99000"
    assert sheet["L3"].value == "99000"


def test_build_export_xlsx_skips_unselected_and_zero_price() -> None:
    submission = Submission(id=9, store_name="فروشگاه تست", shop_id="411147")
    unselected = SubmissionRow(input_row=2, input_product_name="رب", input_price="100000")
    zero_row = SubmissionRow(input_row=3, input_product_name="بالم", input_price="0")
    match = TorobMatch(base_prk="zero", name="بالم", price=80000, rank=0)
    zero_row.selections = [SubmissionSelection(match=match, final_price="0")]

    content = build_export_xlsx(submission, [unselected, zero_row])
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active

    assert sheet.max_row == 1
