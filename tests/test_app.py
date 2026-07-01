from datetime import datetime, timezone
from io import BytesIO
from types import SimpleNamespace

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.database import Base, get_db
from app.main import create_app
from app.models import Store, Submission, SubmissionBatch, SubmissionBatchItem, SubmissionRow, SubmissionSelection, TorobMatch
from app.services.product_search import ProductSearchError, ProductSearchResult
from app.services.torob import TorobClientError, TorobSearchResult
from app.services.uniom import UniomClientError


def make_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["product_name", "price"])
    sheet.append(["رب گوجه روژین", "165000"])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def make_multi_row_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["product_name", "price"])
    sheet.append(["رب گوجه روژین", "165000"])
    sheet.append(["بالم لب", "90000"])
    sheet.append(["شامپو", "120000"])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class FakeProductSearchClient:
    async def search_products(self, query: str, page: int = 0, per_source: int = 2) -> list[ProductSearchResult]:
        return [
            ProductSearchResult(
                source="torob",
                rank=0,
                base_prk="base-1",
                name=f"{query} ترب",
                price=150000,
                price_text="از ۱۵۰۰۰۰ تومان",
                image_url="https://image.example/a.jpg",
                product_url="https://torob.com/p/base-1",
                is_already_added=False,
            ),
            ProductSearchResult(
                source="torob",
                rank=1,
                base_prk="base-2",
                name=f"{query} ترب دوم",
                price=155000,
                price_text="از ۱۵۵۰۰۰ تومان",
                image_url="https://image.example/b.jpg",
                product_url="https://torob.com/p/base-2",
                is_already_added=False,
            ),
            ProductSearchResult(
                source="torob",
                rank=2,
                base_prk="base-3",
                name=f"{query} ترب سوم",
                price=160000,
                price_text="از ۱۶۰۰۰۰ تومان",
                image_url="https://image.example/c.jpg",
                product_url="https://torob.com/p/base-3",
                is_already_added=False,
            ),
            ProductSearchResult(
                source="torob",
                rank=3,
                base_prk="base-4",
                name=f"{query} ترب چهارم",
                price=165000,
                price_text="از ۱۶۵۰۰۰ تومان",
                image_url="https://image.example/d.jpg",
                product_url="https://torob.com/p/base-4",
                is_already_added=False,
            )
        ]

    async def close(self) -> None:
        return None


class FakeForbiddenProductSearchClient:
    async def search_products(self, query: str, page: int = 0, per_source: int = 2) -> list[ProductSearchResult]:
        raise ProductSearchError("torob_forbidden", "اتصال به ترب مجاز نیست.")

    async def close(self) -> None:
        return None


class FakeBotChallengeProductSearchClient:
    async def search_products(self, query: str, page: int = 0, per_source: int = 2) -> list[ProductSearchResult]:
        raise ProductSearchError(
            "torob_bot_challenge",
            "ترب فعلا درخواست‌های جستجوی خودکار را تایید نمی‌کند.",
        )

    async def close(self) -> None:
        return None


class FakeTimeoutProductSearchClient:
    async def search_products(self, query: str, page: int = 0, per_source: int = 2) -> list[ProductSearchResult]:
        raise ProductSearchError("torob_timeout", "جستجو کامل نشد. دوباره تلاش کن.")

    async def close(self) -> None:
        return None


class FakeTimeoutTorobClient:
    async def search_base_products(self, query: str, size: int = 5, page: int = 0) -> list[TorobSearchResult]:
        raise TorobClientError("torob_timeout", "جستجو کامل نشد. دوباره تلاش کن.")

    async def close(self) -> None:
        return None


class FakeChallengeAfterOneSuccessTorobClient:
    def __init__(self) -> None:
        self.calls = 0

    async def search_products(self, query: str, page: int = 0, per_source: int = 2) -> list[ProductSearchResult]:
        self.calls += 1
        if self.calls > 1:
            raise ProductSearchError("torob_bot_challenge", "gateway challenge")
        return [
            ProductSearchResult(
                source="torob",
                rank=0,
                base_prk="base-1",
                name=f"{query} ترب",
                price=150000,
                price_text="از ۱۵۰۰۰۰ تومان",
                image_url="https://image.example/a.jpg",
                product_url="https://torob.com/p/base-1",
                is_already_added=False,
            )
        ]

    async def close(self) -> None:
        return None


class FakeGatewayNotFoundAfterOneSuccessTorobClient:
    def __init__(self) -> None:
        self.calls = 0

    async def search_products(self, query: str, page: int = 0, per_source: int = 2) -> list[ProductSearchResult]:
        self.calls += 1
        if self.calls > 1:
            raise ProductSearchError("torob_gateway_not_found", "مسیر gateway ترب پیدا نشد.")
        return [
            ProductSearchResult(
                source="torob",
                rank=0,
                base_prk="base-1",
                name=f"{query} ترب",
                price=150000,
                price_text="از ۱۵۰۰۰۰ تومان",
                image_url="https://image.example/a.jpg",
                product_url="https://torob.com/p/base-1",
                is_already_added=False,
            )
        ]

    async def close(self) -> None:
        return None


class FakePagedProductSearchClient:
    async def search_products(self, query: str, page: int = 0, per_source: int = 2) -> list[ProductSearchResult]:
        return [
            ProductSearchResult(
                source="torob",
                rank=0,
                base_prk=f"base-{page}-1",
                name=f"{query} ترب {page}-1",
                price=100000 + page,
                price_text=None,
                image_url=None,
                product_url=f"https://torob.com/p/base-{page}-1",
            ),
            ProductSearchResult(
                source="basalam",
                rank=1,
                base_prk=f"basalam-{page}-1",
                name=f"{query} باسلام {page}-1",
                price=110000 + page,
                price_text=None,
                image_url=None,
                product_url=f"https://basalam.com/p/basalam-{page}-1",
            ),
            ProductSearchResult(
                source="torob",
                rank=2,
                base_prk=f"base-{page}-2",
                name=f"{query} ترب {page}-2",
                price=120000 + page,
                price_text=None,
                image_url=None,
                product_url=f"https://torob.com/p/base-{page}-2",
            ),
            ProductSearchResult(
                source="basalam",
                rank=3,
                base_prk=f"basalam-{page}-2",
                name=f"{query} باسلام {page}-2",
                price=130000 + page,
                price_text=None,
                image_url=None,
                product_url=f"https://basalam.com/p/basalam-{page}-2",
            ),
        ]

    async def close(self) -> None:
        return None


class FakeTorobBulkAddClient:
    calls = []

    async def bulk_add(self, shop_id: int, items):
        self.__class__.calls.append((shop_id, items))
        return SimpleNamespace(sent_count=len(items), response_text='[{"ok": true}]')

    async def close(self) -> None:
        return None


class FakeReachableTorobBulkHealthClient:
    async def health_check(self):
        return SimpleNamespace(
            status_code=400,
            outcome="reachable",
            detail="endpoint bulk-add بدون چالش ربات پاسخ داد.",
        )

    async def close(self) -> None:
        return None


class FakeUniomClient:
    async def get_chat(self, chat_id: str) -> dict:
        return {"id": 1, "type": "channel", "username": chat_id.removeprefix("@")}

    async def get_chat_history(self, chat_id: str, limit: int, offset_id: int | None = None) -> list[dict]:
        return [
            {
                "message_id": 11,
                "date": 1755439679,
                "text": "کتونی نایک وودو\r\nقیمت: 668٬000 تومان\r\nارتباط: @seller",
            },
            {
                "message_id": 10,
                "date": 1755439656,
                "photo": [{"file_id": "photo-large", "width": 800, "height": 800}],
            },
        ]

    async def get_chat_history_paginated(self, chat_id: str, total_limit: int, page_size: int) -> list[dict]:
        return await self.get_chat_history(chat_id, total_limit)

    async def get_file(self, file_id: str):
        return SimpleNamespace(file_id=file_id, file_path="files/photo-large.jpg")

    async def download_file(self, file_path: str) -> bytes:
        return b"fake-image"

    async def close(self) -> None:
        return None


class FakeEitaaTorobClient:
    async def search_base_products(self, query: str, size: int = 5, page: int = 0) -> list[TorobSearchResult]:
        return [
            TorobSearchResult(
                rank=0,
                base_prk="nike-voodoo",
                name="کتونی نایک مدل وودو",
                price=668000,
                price_text="از ۶۶۸٬۰۰۰ تومان",
                image_url="https://image.example/nike.jpg",
                product_url="https://torob.com/p/nike-voodoo",
                is_already_added=False,
            )
        ]

    async def search_by_image_bytes(self, image_bytes: bytes, size: int = 5) -> list[TorobSearchResult]:
        return []

    async def close(self) -> None:
        return None


def make_torob_result(base_prk: str, name: str, price: int = 100000) -> TorobSearchResult:
    return TorobSearchResult(
        rank=0,
        base_prk=base_prk,
        name=name,
        price=price,
        price_text=None,
        image_url=None,
        product_url=f"https://torob.com/p/{base_prk}",
        is_already_added=False,
    )


def test_upload_confirm_admin_export(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    monkeypatch.setattr("app.routes.seller.ProductSearchClient", FakeProductSearchClient)
    monkeypatch.setattr("app.routes.seller.SessionLocal", TestingSessionLocal)
    monkeypatch.setattr("app.routes.seller.settings.upload_dir", str(tmp_path / "uploads"))
    monkeypatch.setattr("app.routes.admin.settings.admin_password", "secret")
    monkeypatch.setattr("app.routes.admin.settings.session_secret", "test-cookie")
    monkeypatch.setattr("app.routes.admin.settings.torob_bulk_add_key", "bulk-test-key")
    monkeypatch.setattr("app.routes.admin.TorobBulkAddClient", FakeTorobBulkAddClient)
    FakeTorobBulkAddClient.calls = []

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    excel_page = client.get("/excel")
    assert excel_page.status_code == 200
    assert "دانلود قالب اکسل" not in excel_page.text
    assert "خروجی اکسل نرم‌افزار حسابداری را هم می‌تونی همین‌جا وارد کنی." in excel_page.text

    eitaa_page = client.get("/eitaa")
    assert eitaa_page.status_code == 200
    assert "اگر پیام عکس داشته باشد" not in eitaa_page.text

    upload = client.post(
        "/uploads",
        data={"store_name": "فروشگاه تست", "seller_phone": "09121234567"},
        files={"file": ("products.xlsx", make_xlsx(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload.status_code == 200
    assert "در حال آماده‌سازی محصولات" in upload.text
    assert "fetch(statusUrl" in upload.text

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission is not None
        assert submission.seller_phone == "09121234567"
        assert submission.original_filename == "products.xlsx"
        assert submission.stored_file_path is not None
        assert (tmp_path / "uploads" / f"{submission.id}-products.xlsx").exists()
        row = submission.rows[0]
        match_id = row.matches[0].id
        second_match_id = row.matches[1].id
        submission_id = submission.id
        row_id = row.id

    status = client.get(f"/submissions/{submission_id}/processing-status")
    assert status.status_code == 204
    assert status.headers["HX-Redirect"] == f"/submissions/{submission_id}/match"

    match_page = client.get(f"/submissions/{submission_id}/match")
    assert match_page.status_code == 200
    assert "ردیف 1: رب گوجه" in match_page.text
    assert "ردیف 2: رب گوجه" not in match_page.text
    assert "رب گوجه روژین ترب" in match_page.text
    assert f'value="{match_id}" checked' not in match_page.text
    assert f'value="{second_match_id}" checked' not in match_page.text
    assert "قیمت ترب: 150,000 تومان" in match_page.text
    assert "قیمت ترب: 155,000 تومان" in match_page.text
    assert "قیمت باسلام" not in match_page.text
    assert "بیشتر" in match_page.text
    assert 'value="165,000"' in match_page.text
    assert "formatPriceInput" in match_page.text
    assert "محصول انتخاب شده" in match_page.text

    confirm = client.post(
        f"/submissions/{submission_id}/confirm",
        data={
            f"selected_{row_id}": [str(match_id), str(second_match_id)],
            f"price_{row_id}": "170000",
            "price_unit": "toman",
        },
    )
    assert confirm.status_code == 200
    assert "کد پیگیری" in confirm.text

    login = client.post("/admin/login", data={"password": "secret"}, follow_redirects=False)
    assert login.status_code == 303

    update = client.post(
        f"/admin/submissions/{submission_id}/shop-id",
        data={"shop_id": "411147"},
        follow_redirects=False,
    )
    assert update.status_code == 303

    export = client.get(f"/admin/submissions/{submission_id}/export.xlsx")
    assert export.status_code == 200
    workbook = load_workbook(BytesIO(export.content))
    sheet = workbook.active
    assert sheet["B2"].value == "فروشگاه تست"
    assert sheet["C2"].value == "09121234567"
    assert sheet["D2"].value == "411147"
    assert sheet["H2"].value == "torob"
    assert sheet["I2"].value == "base-1"
    assert sheet["L2"].value == "170000"
    assert sheet["H3"].value == "torob"
    assert sheet["I3"].value == "base-2"
    assert sheet["L3"].value == "170000"

    detail = client.get(f"/admin/submissions/{submission_id}")
    assert detail.status_code == 200
    assert "09121234567" in detail.text

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        batch_id = submission.batches[0].id

    send = client.post(
        f"/admin/submissions/{submission_id}/batches/{batch_id}/send-torob",
        follow_redirects=False,
    )
    assert send.status_code == 303
    assert FakeTorobBulkAddClient.calls
    sent_shop_id, sent_items = FakeTorobBulkAddClient.calls[0]
    assert sent_shop_id == 411147
    assert len(sent_items) == 2
    assert sent_items[0].base_product_rk == "base-1"
    assert sent_items[0].price == 170000
    assert sent_items[1].base_product_rk == "base-2"
    assert sent_items[1].price == 170000

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        batch = submission.batches[0]
        assert batch.status == "sent"
        assert batch.torob_sent_count == 2
        assert batch.torob_skipped_count == 0

    listing = client.get("/admin/submissions")
    assert listing.status_code == 200
    assert "09121234567" in listing.text


def test_forbidden_torob_connection_fails_job_with_clean_message(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    monkeypatch.setattr("app.routes.seller.ProductSearchClient", FakeForbiddenProductSearchClient)
    monkeypatch.setattr("app.routes.seller.SessionLocal", TestingSessionLocal)
    monkeypatch.setattr("app.routes.seller.settings.upload_dir", str(tmp_path / "uploads"))

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    upload = client.post(
        "/uploads",
        data={"store_name": "فروشگاه تست", "seller_phone": ""},
        files={"file": ("products.xlsx", make_xlsx(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload.status_code == 200

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission is not None
        assert submission.status == "failed"
        submission_id = submission.id

    status = client.get(f"/submissions/{submission_id}/processing-status")
    assert status.status_code == 200
    assert status.headers["X-Processing-State"] == "failed"
    assert "اتصال به ترب مجاز نیست" in status.text
    assert "api.torob.com" not in status.text


def test_bot_challenge_fails_job_instead_of_row_errors(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    monkeypatch.setattr("app.routes.seller.ProductSearchClient", FakeBotChallengeProductSearchClient)
    monkeypatch.setattr("app.routes.seller.SessionLocal", TestingSessionLocal)
    monkeypatch.setattr("app.routes.seller.settings.upload_dir", str(tmp_path / "uploads"))

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    upload = client.post(
        "/uploads",
        data={"store_name": "فروشگاه تست", "seller_phone": ""},
        files={"file": ("products.xlsx", make_xlsx(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload.status_code == 200

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission is not None
        assert submission.status == "failed"
        assert submission.rows[0].error_message is None
        submission_id = submission.id

    status = client.get(f"/submissions/{submission_id}/processing-status")
    assert status.status_code == 200
    assert status.headers["X-Processing-State"] == "failed"
    assert "صفحه بررسی ربات" in status.text
    assert "فایل اکسل نیست" in status.text


def test_timeout_before_first_success_fails_job_with_vpn_hint(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    monkeypatch.setattr("app.routes.seller.ProductSearchClient", FakeTimeoutProductSearchClient)
    monkeypatch.setattr("app.routes.seller.SessionLocal", TestingSessionLocal)
    monkeypatch.setattr("app.routes.seller.settings.upload_dir", str(tmp_path / "uploads"))

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    upload = client.post(
        "/uploads",
        data={"store_name": "فروشگاه تست", "seller_phone": ""},
        files={"file": ("products.xlsx", make_xlsx(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload.status_code == 200

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission is not None
        assert submission.status == "failed"
        assert submission.rows[0].error_message is None
        submission_id = submission.id

    status = client.get(f"/submissions/{submission_id}/processing-status")
    assert status.status_code == 200
    assert status.headers["X-Processing-State"] == "failed"
    assert "VPN" in status.text
    assert "فایل اکسل نیست" in status.text


def test_gateway_challenge_after_success_keeps_partial_results(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    monkeypatch.setattr("app.routes.seller.ProductSearchClient", FakeChallengeAfterOneSuccessTorobClient)
    monkeypatch.setattr("app.routes.seller.SessionLocal", TestingSessionLocal)
    monkeypatch.setattr("app.routes.seller.settings.upload_dir", str(tmp_path / "uploads"))

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    upload = client.post(
        "/uploads",
        data={"store_name": "فروشگاه تست", "seller_phone": ""},
        files={"file": ("products.xlsx", make_multi_row_xlsx(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload.status_code == 200

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission is not None
        assert submission.status == "ready"
        assert submission.rows[0].matches
        assert "ترب کامل نشد" in submission.rows[1].error_message
        assert "کمی بعد تلاش مجدد" in submission.rows[2].error_message
        submission_id = submission.id

    status = client.get(f"/submissions/{submission_id}/processing-status")
    assert status.status_code == 204
    assert status.headers["HX-Redirect"] == f"/submissions/{submission_id}/match"


def test_gateway_404_after_success_keeps_partial_results(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    monkeypatch.setattr("app.routes.seller.ProductSearchClient", FakeGatewayNotFoundAfterOneSuccessTorobClient)
    monkeypatch.setattr("app.routes.seller.SessionLocal", TestingSessionLocal)
    monkeypatch.setattr("app.routes.seller.settings.upload_dir", str(tmp_path / "uploads"))

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    upload = client.post(
        "/uploads",
        data={"store_name": "فروشگاه تست", "seller_phone": ""},
        files={"file": ("products.xlsx", make_multi_row_xlsx(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload.status_code == 200

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission is not None
        assert submission.status == "ready"
        assert submission.rows[0].matches
        assert "gateway" in submission.rows[1].error_message
        assert "کمی بعد تلاش مجدد" in submission.rows[2].error_message
        submission_id = submission.id

    status = client.get(f"/submissions/{submission_id}/processing-status")
    assert status.status_code == 204
    assert status.headers["HX-Redirect"] == f"/submissions/{submission_id}/match"


def test_admin_torob_health_reports_timeout(monkeypatch) -> None:
    monkeypatch.setattr("app.routes.admin.TorobClient", FakeTimeoutTorobClient)
    monkeypatch.setattr("app.routes.admin.settings.admin_password", "secret")
    monkeypatch.setattr("app.routes.admin.settings.session_secret", "test-cookie")

    app = create_app()
    client = TestClient(app)

    login = client.post("/admin/login", data={"password": "secret"}, follow_redirects=False)
    assert login.status_code == 303

    health = client.get("/admin/torob-health")

    assert health.status_code == 503
    assert "torob_timeout" in health.text
    assert "VPN" in health.text


def test_admin_torob_bulk_health_reports_reachable(monkeypatch) -> None:
    monkeypatch.setattr("app.routes.admin.TorobBulkAddClient", FakeReachableTorobBulkHealthClient)
    monkeypatch.setattr("app.routes.admin.settings.admin_password", "secret")
    monkeypatch.setattr("app.routes.admin.settings.session_secret", "test-cookie")
    monkeypatch.setattr("app.routes.admin.settings.torob_bulk_add_key", "bulk-key")
    monkeypatch.setattr("app.routes.admin.settings.torob_iw1_header", "iw1")

    app = create_app()
    client = TestClient(app)

    login = client.post("/admin/login", data={"password": "secret"}, follow_redirects=False)
    assert login.status_code == 303

    health = client.get("/admin/torob-bulk-health")

    assert health.status_code == 200
    assert "OK" in health.text
    assert "TOROB_IW1_HEADER=set" in health.text
    assert "TOROB_BULK_ADD_KEY=set" in health.text


def test_retry_row_search_replaces_error_with_results(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    monkeypatch.setattr("app.routes.seller.ProductSearchClient", FakeProductSearchClient)

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    with TestingSessionLocal() as db:
        submission = Submission(store_name="فروشگاه تست", status="ready", total_rows=1)
        db.add(submission)
        db.commit()
        db.refresh(submission)
        row = SubmissionRow(
            submission_id=submission.id,
            input_row=1,
            input_product_name="رب گوجه روژین",
            input_price="165000",
            error_message="جستجو کامل نشد. دوباره تلاش کن.",
        )
        db.add(row)
        db.commit()
        row_id = row.id

    retry = client.post(f"/rows/{row_id}/retry")

    assert retry.status_code == 200
    assert "رب گوجه روژین ترب" in retry.text
    assert "تلاش مجدد" not in retry.text


def test_load_more_row_matches_appends_next_batch(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    monkeypatch.setattr("app.routes.seller.ProductSearchClient", FakePagedProductSearchClient)

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    with TestingSessionLocal() as db:
        submission = Submission(store_name="فروشگاه تست", status="ready", total_rows=1)
        db.add(submission)
        db.commit()
        db.refresh(submission)
        row = SubmissionRow(
            submission_id=submission.id,
            input_row=1,
            input_product_name="رب",
            has_more_matches=True,
            next_search_page=1,
        )
        row.matches = [
            TorobMatch(source="torob", base_prk="base-0-1", name="قبلی ۱", price=1, rank=0),
            TorobMatch(source="basalam", base_prk="basalam-0-1", name="قبلی ۲", price=2, rank=1),
            TorobMatch(source="torob", base_prk="base-0-2", name="قبلی ۳", price=3, rank=2),
            TorobMatch(source="basalam", base_prk="basalam-0-2", name="قبلی ۴", price=4, rank=3),
        ]
        db.add(row)
        db.commit()
        row_id = row.id

    response = client.post(f"/rows/{row_id}/more")

    assert response.status_code == 200
    assert "رب ترب 1-1" in response.text
    assert "price-unit-card" in response.text
    with TestingSessionLocal() as db:
        row = db.query(SubmissionRow).filter(SubmissionRow.id == row_id).first()
        assert len(row.matches) == 8
        assert row.next_search_page == 2


def test_confirm_converts_rial_prices_to_toman(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    with TestingSessionLocal() as db:
        submission = Submission(store_name="فروشگاه تست", status="ready", total_rows=1)
        db.add(submission)
        db.commit()
        db.refresh(submission)
        row = SubmissionRow(submission_id=submission.id, input_row=1, input_product_name="رب", input_price=None)
        db.add(row)
        db.commit()
        db.refresh(row)
        row.matches.append(
            TorobMatch(
                source="torob",
                base_prk="base-rial",
                name="رب ترب",
                price=120000,
                rank=0,
            )
        )
        db.commit()
        submission_id = submission.id
        row_id = row.id
        match_id = row.matches[0].id

    confirm = client.post(
        f"/submissions/{submission_id}/confirm",
        data={f"selected_{row_id}": str(match_id), f"price_{row_id}": "1200000", "price_unit": "rial"},
    )
    assert confirm.status_code == 200

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission.rows[0].final_price == "120000"


def test_continue_submission_hides_submitted_rows_and_shows_resume_card(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    monkeypatch.setattr("app.routes.seller.ProductSearchClient", FakeProductSearchClient)
    monkeypatch.setattr("app.routes.seller.SessionLocal", TestingSessionLocal)
    monkeypatch.setattr("app.routes.seller.settings.upload_dir", str(tmp_path / "uploads"))
    submitted_times = iter(
        [
            datetime(2026, 6, 7, 10, 0, tzinfo=timezone.utc),
            datetime(2026, 6, 7, 10, 5, tzinfo=timezone.utc),
        ]
    )
    monkeypatch.setattr("app.routes.seller.utc_now", lambda: next(submitted_times))
    monkeypatch.setattr("app.routes.admin.settings.admin_password", "secret")
    monkeypatch.setattr("app.routes.admin.settings.session_secret", "test-cookie")
    monkeypatch.setattr("app.routes.admin.settings.torob_bulk_add_key", "bulk-test-key")
    monkeypatch.setattr("app.routes.admin.TorobBulkAddClient", FakeTorobBulkAddClient)
    FakeTorobBulkAddClient.calls = []

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    upload = client.post(
        "/uploads",
        data={"store_name": "فروشگاه تست", "seller_phone": ""},
        files={"file": ("products.xlsx", make_multi_row_xlsx(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload.status_code == 200

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission is not None
        first_row = submission.rows[0]
        second_row = submission.rows[1]
        submission_id = submission.id
        first_row_id = first_row.id
        first_match_id = first_row.matches[0].id
        second_row_id = second_row.id
        second_match_id = second_row.matches[0].id

    confirm = client.post(
        f"/submissions/{submission_id}/confirm",
        data={
            f"selected_{first_row_id}": str(first_match_id),
            f"price_{first_row_id}": "170000",
            "price_unit": "toman",
            "finish_mode": "continue",
        },
    )
    assert confirm.status_code == 200
    assert "1 کالا ثبت شد و به زودی در فروشگاهت قرار می‌گیرد" in confirm.text
    assert "مشتری بیشتری می‌گیری" in confirm.text

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission.status == "ready"
        assert submission.rows[0].submitted_at is not None
        assert submission.rows[1].submitted_at is None

    match_page = client.get(f"/submissions/{submission_id}/match")
    assert match_page.status_code == 200
    assert "رب گوجه روژین ترب" not in match_page.text
    assert "بالم لب ترب" in match_page.text

    draft = client.post(
        f"/submissions/{submission_id}/draft",
        data={
            f"selected_{second_row_id}": str(second_match_id),
            f"price_{second_row_id}": "99000",
            "price_unit": "toman",
        },
    )
    assert draft.status_code == 200

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert len(submission.rows[0].selections) == 1
        assert len(submission.rows[1].selections) == 1

    home = client.get("/")
    assert home.status_code == 200
    assert "انتخاب محصولات فروشگاه تست ناتمام مانده" not in home.text
    assert 'href="/excel"' in home.text

    excel_page = client.get("/excel")
    assert excel_page.status_code == 200
    assert "انتخاب محصولات فروشگاه تست ناتمام مانده" in excel_page.text
    assert "2 ردیف هنوز مانده" in excel_page.text

    second_confirm = client.post(
        f"/submissions/{submission_id}/confirm",
        data={
            f"selected_{second_row_id}": str(second_match_id),
            f"price_{second_row_id}": "99000",
            "price_unit": "toman",
            "finish_mode": "continue",
        },
    )
    assert second_confirm.status_code == 200

    login = client.post("/admin/login", data={"password": "secret"}, follow_redirects=False)
    assert login.status_code == 303

    update_shop_id = client.post(
        f"/admin/submissions/{submission_id}/shop-id",
        data={"shop_id": "411147"},
        follow_redirects=False,
    )
    assert update_shop_id.status_code == 303

    detail = client.get(f"/admin/submissions/{submission_id}")
    assert detail.status_code == 200
    assert "ثبت‌های مرحله‌ای" in detail.text
    assert detail.text.count(f"/admin/submissions/{submission_id}/export.xlsx?batch=") == 2

    all_export = client.get(f"/admin/submissions/{submission_id}/export.xlsx")
    all_workbook = load_workbook(BytesIO(all_export.content))
    assert all_workbook.active.max_row == 3

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert len(submission.batches) == 2
        first_batch = [batch for batch in submission.batches if batch.items[0].row_id == first_row_id][0]
        second_batch = [batch for batch in submission.batches if batch.items[0].row_id == second_row_id][0]
        assert first_batch.selected_count == 1
        assert second_batch.selected_count == 1

    first_export = client.get(f"/admin/submissions/{submission_id}/export.xlsx?batch={first_batch.id}")
    first_workbook = load_workbook(BytesIO(first_export.content))
    assert first_workbook.active.max_row == 2
    assert first_workbook.active["F2"].value == "رب گوجه روژین"

    second_export = client.get(f"/admin/submissions/{submission_id}/export.xlsx?batch={second_batch.id}")
    second_workbook = load_workbook(BytesIO(second_export.content))
    assert second_workbook.active.max_row == 2
    assert second_workbook.active["F2"].value == "بالم لب"

    send = client.post(
        f"/admin/submissions/{submission_id}/batches/{first_batch.id}/send-torob",
        follow_redirects=False,
    )
    assert send.status_code == 303
    assert FakeTorobBulkAddClient.calls
    sent_shop_id, sent_items = FakeTorobBulkAddClient.calls[0]
    assert sent_shop_id == 411147
    assert sent_items[0].base_product_rk == "base-1"
    assert sent_items[0].price == 170000

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        sent_batch = [batch for batch in submission.batches if batch.id == first_batch.id][0]
        assert sent_batch.status == "sent"
        assert sent_batch.torob_sent_count == 1
        assert sent_batch.torob_skipped_count == 0

    detail_after_send = client.get(f"/admin/submissions/{submission_id}")
    assert "ارسال شد: 1" in detail_after_send.text


def test_admin_can_rebuild_batch_from_legacy_selections(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    monkeypatch.setattr("app.routes.admin.settings.admin_password", "secret")
    monkeypatch.setattr("app.routes.admin.settings.session_secret", "test-cookie")
    monkeypatch.setattr("app.routes.admin.settings.torob_bulk_add_key", "bulk-test-key")
    monkeypatch.setattr("app.routes.admin.TorobBulkAddClient", FakeTorobBulkAddClient)
    FakeTorobBulkAddClient.calls = []

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    created_at = datetime(2026, 6, 8, 10, 0, tzinfo=timezone.utc)
    with TestingSessionLocal() as db:
        submission = Submission(
            store_name="فروشگاه قدیمی",
            seller_phone="09900900375",
            shop_id="411488",
            status="submitted",
            total_rows=1,
            selected_rows=1,
            created_at=created_at,
            updated_at=created_at,
        )
        db.add(submission)
        db.commit()
        db.refresh(submission)
        row = SubmissionRow(
            submission_id=submission.id,
            input_row=1,
            input_product_name="رب گوجه",
            input_price="100000",
            final_price="100000",
            submitted_at=created_at,
        )
        db.add(row)
        db.commit()
        db.refresh(row)
        match = TorobMatch(
            row_id=row.id,
            source="torob",
            rank=0,
            base_prk="legacy-base",
            name="رب گوجه تست",
            price=95000,
            price_text=None,
            image_url=None,
            product_url=None,
            is_already_added=False,
        )
        db.add(match)
        db.commit()
        db.refresh(match)
        row.selected_match_id = match.id
        db.add(SubmissionSelection(row_id=row.id, match_id=match.id, final_price="100000", created_at=created_at))
        db.commit()
        submission_id = submission.id

    login = client.post("/admin/login", data={"password": "secret"}, follow_redirects=False)
    assert login.status_code == 303

    detail = client.get(f"/admin/submissions/{submission_id}")
    assert detail.status_code == 200
    assert "ساخت ثبت ارسال از انتخاب‌های قبلی" in detail.text

    rebuild = client.post(f"/admin/submissions/{submission_id}/rebuild-batch", follow_redirects=False)
    assert rebuild.status_code == 303

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert len(submission.batches) == 1
        assert submission.batches[0].selected_count == 1
        batch_id = submission.batches[0].id

    detail_after = client.get(f"/admin/submissions/{submission_id}")
    assert detail_after.status_code == 200
    assert "ارسال به ترب" in detail_after.text

    send = client.post(
        f"/admin/submissions/{submission_id}/batches/{batch_id}/send-torob",
        follow_redirects=False,
    )
    assert send.status_code == 303
    sent_shop_id, sent_items = FakeTorobBulkAddClient.calls[0]
    assert sent_shop_id == 411488
    assert sent_items[0].base_product_rk == "legacy-base"
    assert sent_items[0].price == 100000


def test_eitaa_import_auto_matches_and_waits_for_seller_preview(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    monkeypatch.setattr("app.routes.seller.SessionLocal", TestingSessionLocal)
    monkeypatch.setattr("app.routes.seller.UniomClient", FakeUniomClient)
    monkeypatch.setattr("app.routes.seller.TorobClient", FakeEitaaTorobClient)
    monkeypatch.setattr("app.routes.seller.settings.uniom_bot_token", "test-token")
    monkeypatch.setattr("app.routes.seller.settings.upload_dir", str(tmp_path / "uploads"))
    monkeypatch.setattr("app.routes.admin.settings.admin_password", "secret")
    monkeypatch.setattr("app.routes.admin.settings.session_secret", "test-cookie")
    monkeypatch.setattr("app.routes.admin.settings.torob_bulk_add_key", "bulk-test-key")

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    response = client.post(
        "/eitaa/import",
        data={"store_name": "فروشگاه ایتا", "seller_phone": "09121234567", "channel_id": "regaal"},
    )
    assert response.status_code == 200
    assert "کانال دریافت شد" in response.text

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission is not None
        assert submission.source == "eitaa"
        assert submission.source_ref == "@regaal"
        assert submission.operation == "add"
        assert submission.store is not None
        assert submission.store.eitaa_channel_id == "@regaal"
        assert submission.store.name == submission.store_name
        assert submission.status == "ready"
        assert submission.price_unit == "toman"
        assert submission.total_rows == 1
        assert submission.selected_rows == 1
        assert len(submission.rows) == 1
        row = submission.rows[0]
        assert row.input_product_name == "کتونی نایک وودو"
        assert row.input_price == "668000"
        assert row.final_price == "668000"
        assert row.submitted_at is None
        assert row.source_image_path is not None
        assert row.selected_match.base_prk == "nike-voodoo"
        assert row.auto_match_score >= 72
        assert len(row.selections) == 1
        assert len(submission.batches) == 0
        submission_id = submission.id
        row_id = row.id
        match_id = row.matches[0].id

    status = client.get(f"/submissions/{submission_id}/processing-status")
    assert status.status_code == 204
    assert status.headers["HX-Redirect"] == f"/submissions/{submission_id}/match"

    match_page = client.get(f"/submissions/{submission_id}/match")
    assert match_page.status_code == 200
    assert "بررسی نهایی محصولات کانال" in match_page.text
    assert "کتونی نایک مدل وودو" in match_page.text
    assert f'value="{match_id}" checked' in match_page.text
    assert 'value="668,000"' in match_page.text
    assert "قیمت‌ها ریال هست یا تومان؟" not in match_page.text
    assert "تایید و ثبت" in match_page.text

    confirm = client.post(
        f"/submissions/{submission_id}/confirm",
        data={
            f"selected_{row_id}": str(match_id),
            f"price_{row_id}": "670000",
            "price_unit": "toman",
        },
    )
    assert confirm.status_code == 200
    assert "محصولات ثبت شد" in confirm.text
    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission.status == "submitted"
        assert submission.rows[0].submitted_at is not None
        assert submission.rows[0].final_price == "670000"
        assert len(submission.batches) == 1
        assert submission.batches[0].selected_count == 1

    login = client.post("/admin/login", data={"password": "secret"}, follow_redirects=False)
    assert login.status_code == 303
    update_shop_id = client.post(
        f"/admin/submissions/{submission_id}/shop-id",
        data={"shop_id": "411488"},
        follow_redirects=False,
    )
    assert update_shop_id.status_code == 303
    detail = client.get(f"/admin/submissions/{submission_id}")
    assert detail.status_code == 200
    assert "@regaal" in detail.text
    assert "ارسال به ترب" in detail.text
    assert "nike-voodoo" not in detail.text


def test_admin_can_create_eitaa_update_preview_from_uniom_updates(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    with TestingSessionLocal() as db:
        store = Store(
            name="Test Store",
            seller_phone="09120000000",
            shop_id="411488",
            eitaa_channel_id="@regaal",
            eitaa_last_update_id=84,
        )
        submission = Submission(
            store=store,
            store_name="Test Store",
            seller_phone="09120000000",
            shop_id="411488",
            source="eitaa",
            source_ref="@regaal",
            operation="add",
            status="submitted",
            price_unit="toman",
        )
        row = SubmissionRow(
            submission=submission,
            input_row=1,
            input_product_name="کفش تست",
            input_price="650000",
            final_price="650000",
            source_message_id="124",
        )
        match = TorobMatch(
            row=row,
            source="torob",
            rank=0,
            base_prk="shoe-base-rk",
            name="کفش تست ترب",
            price=650000,
        )
        db.add_all([store, submission, row, match])
        db.flush()
        row.selected_match_id = match.id
        db.add(SubmissionSelection(row_id=row.id, match_id=match.id, final_price="650000"))
        db.commit()
        submission_id = submission.id

    class UpdateUniomClient:
        calls = []

        async def get_updates(self, offset=None, timeout_seconds=30):
            self.calls.append((offset, timeout_seconds))
            return [
                {
                    "update_id": 85,
                    "edited_channel_post": {
                        "message_id": 124,
                        "date": 1781681754,
                        "edit_date": 1781682789,
                        "chat": {"username": "regaal"},
                        "text": "کفش تست - 750000 تومان",
                    },
                }
            ]

        async def close(self):
            return None

    monkeypatch.setattr("app.routes.admin.UniomClient", UpdateUniomClient)
    monkeypatch.setattr("app.routes.admin.settings.admin_password", "secret")
    monkeypatch.setattr("app.routes.admin.settings.session_secret", "test-cookie")
    monkeypatch.setattr("app.routes.admin.settings.uniom_bot_token", "test-token")

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)
    login = client.post("/admin/login", data={"password": "secret"}, follow_redirects=False)
    assert login.status_code == 303

    response = client.post(f"/admin/submissions/{submission_id}/eitaa-updates/check", follow_redirects=False)

    assert response.status_code == 303
    assert UpdateUniomClient.calls == [(85, 30)]
    with TestingSessionLocal() as db:
        store = db.query(Store).filter(Store.eitaa_channel_id == "@regaal").one()
        assert store.eitaa_last_update_id == 85
        previews = db.query(Submission).filter(Submission.operation == "price_update").all()
        assert len(previews) == 1
        preview = previews[0]
        assert preview.store_id == store.id
        assert preview.source == "eitaa_update"
        assert preview.selected_rows == 1
        assert preview.rows[0].final_price == "750000"
        assert preview.rows[0].selections[0].match.base_prk == "shoe-base-rk"
        assert f"/admin/submissions/{preview.id}" in response.headers["location"]


def test_eitaa_image_download_failure_does_not_fail_processing(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    class ImageFailUniomClient(FakeUniomClient):
        async def download_file(self, file_path: str) -> bytes:
            raise UniomClientError("uniom_file_unavailable", "عکس محصول از ایتا کامل دریافت نشد.")

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    monkeypatch.setattr("app.routes.seller.SessionLocal", TestingSessionLocal)
    monkeypatch.setattr("app.routes.seller.UniomClient", ImageFailUniomClient)
    monkeypatch.setattr("app.routes.seller.TorobClient", FakeEitaaTorobClient)
    monkeypatch.setattr("app.routes.seller.settings.uniom_bot_token", "test-token")
    monkeypatch.setattr("app.routes.seller.settings.upload_dir", str(tmp_path / "uploads"))

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    response = client.post(
        "/eitaa/import",
        data={"store_name": "فروشگاه ایتا", "seller_phone": "", "channel_id": "regaal"},
    )
    assert response.status_code == 200

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission.status == "ready"
        assert submission.error_message is None
        row = submission.rows[0]
        assert row.source_image_path is None
        assert row.matches
        assert row.selections


@pytest.mark.asyncio
async def test_eitaa_text_search_filters_unrelated_short_query_results() -> None:
    from app.routes.seller import _search_eitaa_text_results

    class ProduceTorobClient:
        async def search_base_products(self, query: str, size: int = 5, page: int = 0) -> list[TorobSearchResult]:
            return [
                make_torob_result("papaya", "پاپایا درجه یک بسته ای", 595000),
                make_torob_result("potato", "سیب زمینی درجه یک یک کیلوگرم", 50000),
                make_torob_result("lettuce", "کاهو فرانسوی درجه یک بسته ای", 95000),
                make_torob_result("cucumber", "خیار گلخانه ای ممتاز یک کیلویی", 125000),
            ]

    results = await _search_eitaa_text_results(ProduceTorobClient(), "خیار درجه یک", {})

    assert [item.base_prk for item in results] == ["cucumber"]


@pytest.mark.asyncio
async def test_eitaa_text_search_keeps_related_product_variants() -> None:
    from app.routes.seller import _search_eitaa_text_results

    class CabbageTorobClient:
        async def search_base_products(self, query: str, size: int = 5, page: int = 0) -> list[TorobSearchResult]:
            return [
                make_torob_result("white-cabbage", "کلم سفید درجه یک", 45000),
                make_torob_result("red-cabbage", "کلم قرمز درجه یک", 45000),
                make_torob_result("lettuce", "کاهو رسمی پاک شده", 75000),
            ]

    results = await _search_eitaa_text_results(CabbageTorobClient(), "کلم سفید و قرمز", {})

    assert [item.base_prk for item in results] == ["white-cabbage", "red-cabbage"]


def test_eitaa_candidate_filter_rejects_seed_for_fresh_produce() -> None:
    from app.routes.seller import _is_plausible_eitaa_candidate

    result = make_torob_result("seed", "بذر خیار درختی درجه یک", 100000)

    assert _is_plausible_eitaa_candidate("خیار تازه درجه یک", result) is False


def test_eitaa_candidate_filter_keeps_seed_when_seed_is_requested() -> None:
    from app.routes.seller import _is_plausible_eitaa_candidate

    result = make_torob_result("seed", "بذر خیار درختی درجه یک", 100000)

    assert _is_plausible_eitaa_candidate("بذر خیار بوته ای", result) is True


def test_eitaa_candidate_filter_keeps_cabbage_color_variants() -> None:
    from app.routes.seller import _is_plausible_eitaa_candidate

    white = make_torob_result("white", "کلم سفید درجه یک", 45000)
    red = make_torob_result("red", "کلم قرمز درجه یک", 45000)

    assert _is_plausible_eitaa_candidate("کلم سفید و قرمز درجه یک", white) is True
    assert _is_plausible_eitaa_candidate("کلم سفید و قرمز درجه یک", red) is True


def test_eitaa_candidate_filter_rejects_motor_oil_for_cooking_oil() -> None:
    from app.routes.seller import _is_plausible_eitaa_candidate

    result = make_torob_result("motor-oil", "روغن موتور لادن مدل 10W40", 500000)

    assert _is_plausible_eitaa_candidate("روغن لادن", result) is False


@pytest.mark.asyncio
async def test_eitaa_text_search_skips_broader_query_when_exact_query_has_candidates() -> None:
    from app.routes.seller import _search_eitaa_text_results

    class CountingTorobClient:
        def __init__(self) -> None:
            self.queries: list[str] = []

        async def search_base_products(self, query: str, size: int = 5, page: int = 0) -> list[TorobSearchResult]:
            self.queries.append(query)
            return [make_torob_result("tomato", "رب گوجه روژین", 120000)]

    client = CountingTorobClient()
    results = await _search_eitaa_text_results(client, "رب گوجه روژین ۸۰۰ گرم", {})

    assert [item.base_prk for item in results] == ["tomato"]
    assert client.queries == ["رب گوجه روژین ۸۰۰ گرم"]


@pytest.mark.asyncio
async def test_eitaa_text_search_does_not_run_a_second_query_when_exact_results_are_unrelated() -> None:
    from app.routes.seller import _search_eitaa_text_results

    class CountingTorobClient:
        def __init__(self) -> None:
            self.queries: list[str] = []

        async def search_base_products(self, query: str, size: int = 5, page: int = 0) -> list[TorobSearchResult]:
            self.queries.append(query)
            return [make_torob_result("unrelated", "شامپو مو", 200000)]

    client = CountingTorobClient()
    results = await _search_eitaa_text_results(client, "رب گوجه روژین ۸۰۰ گرم", {})

    assert results == []
    assert client.queries == ["رب گوجه روژین ۸۰۰ گرم"]


def test_eitaa_processing_status_is_indeterminate_before_rows_are_known(tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    with TestingSessionLocal() as db:
        submission = Submission(
            store_name="فروشگاه تست",
            source="eitaa",
            source_ref="@kosarmarket",
            status="processing",
        )
        db.add(submission)
        db.commit()
        submission_id = submission.id

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    status = client.get(f"/submissions/{submission_id}/processing-status")

    assert status.status_code == 200
    assert "در حال خواندن پیام‌های کانال ایتا" in status.text
    assert "0 از 0" not in status.text
    assert "is-indeterminate" in status.text


def test_eitaa_processing_status_uses_known_total_rows(tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    with TestingSessionLocal() as db:
        submission = Submission(
            store_name="فروشگاه تست",
            source="eitaa",
            source_ref="@timanic_shop",
            status="processing",
            total_rows=400,
        )
        db.add(submission)
        db.commit()
        db.refresh(submission)
        for index in range(1, 9):
            row = SubmissionRow(
                submission_id=submission.id,
                input_row=index,
                input_product_name=f"محصول {index}",
            )
            db.add(row)
            db.flush()
            db.add(
                TorobMatch(
                    row_id=row.id,
                    source="torob",
                    rank=0,
                    base_prk=f"base-{index}",
                    name=f"محصول {index}",
                    price=100000,
                    image_url=None,
                    product_url=None,
                    is_already_added=False,
                )
            )
        db.commit()
        submission_id = submission.id

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    status = client.get(f"/submissions/{submission_id}/processing-status")

    assert status.status_code == 200
    assert "8 از 400 ردیف پردازش شده" in status.text
    assert "8 از 8" not in status.text


def test_eitaa_import_keeps_no_price_products_for_seller_review(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    class NoPriceUniomClient(FakeUniomClient):
        async def get_chat_history(self, chat_id: str, limit: int, offset_id: int | None = None) -> list[dict]:
            return [
                {
                    "message_id": 21,
                    "date": 1755439679,
                    "text": "روغن جامد لادن\r\nدر دو نوع طلایی و آبی\r\nبه تعداد محدود شارژ شدن",
                    "photo": [{"file_id": "oil-photo", "width": 900, "height": 900}],
                }
            ]

        async def get_file(self, file_id: str):
            return SimpleNamespace(file_id=file_id, file_path="files/oil.jpg")

    class NoPriceTorobClient(FakeEitaaTorobClient):
        async def search_base_products(self, query: str, size: int = 5, page: int = 0) -> list[TorobSearchResult]:
            return [
                TorobSearchResult(
                    rank=0,
                    base_prk="laden-oil",
                    name="روغن جامد لادن طلایی ۹۰۰ گرم",
                    price=240000,
                    price_text="از ۲۴۰٬۰۰۰ تومان",
                    image_url="https://image.example/oil.jpg",
                    product_url="https://torob.com/p/laden-oil",
                    is_already_added=False,
                )
            ]

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    monkeypatch.setattr("app.routes.seller.SessionLocal", TestingSessionLocal)
    monkeypatch.setattr("app.routes.seller.UniomClient", NoPriceUniomClient)
    monkeypatch.setattr("app.routes.seller.TorobClient", NoPriceTorobClient)
    monkeypatch.setattr("app.routes.seller.settings.uniom_bot_token", "test-token")
    monkeypatch.setattr("app.routes.seller.settings.upload_dir", str(tmp_path / "uploads"))

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    response = client.post(
        "/eitaa/import",
        data={"store_name": "فروشگاه ایتا", "seller_phone": "", "channel_id": "kosarmarket"},
    )
    assert response.status_code == 200

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission is not None
        assert submission.status == "ready"
        assert submission.total_rows == 1
        assert submission.selected_rows == 0
        row = submission.rows[0]
        assert row.input_product_name == "روغن جامد لادن"
        assert row.input_price is None
        assert row.final_price is None
        assert row.error_message is None
        assert row.matches[0].base_prk == "laden-oil"
        submission_id = submission.id
        row_id = row.id
        match_id = row.matches[0].id

    summary = client.get(f"/submissions/{submission_id}/eitaa-summary")
    assert summary.status_code == 200
    assert "1 محصول قیمت نداشتند" in summary.text
    assert "تکمیل محصولات بدون قیمت" in summary.text

    match_page = client.get(f"/submissions/{submission_id}/match")
    assert match_page.status_code == 200
    assert "روغن جامد لادن طلایی ۹۰۰ گرم" in match_page.text
    assert f'name="price_{row_id}"' in match_page.text

    confirm = client.post(
        f"/submissions/{submission_id}/confirm",
        data={f"selected_{row_id}": str(match_id), f"price_{row_id}": "245000", "price_unit": "toman"},
    )
    assert confirm.status_code == 200
    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission.selected_rows == 1
        assert submission.rows[0].final_price == "245000"
        assert len(submission.batches) == 1


def test_eitaa_preview_groups_review_rows_and_supports_continue(monkeypatch, tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    with TestingSessionLocal() as db:
        submission = Submission(
            store_name="فروشگاه ایتا",
            source="eitaa",
            source_ref="@kosarmarket",
            status="ready",
            price_unit="toman",
            total_rows=2,
            selected_rows=1,
        )
        db.add(submission)
        db.commit()
        db.refresh(submission)
        ready_row = SubmissionRow(
            submission_id=submission.id,
            input_row=1,
            input_product_name="روغن جامد لادن",
            input_price="330000",
            final_price="330000",
        )
        review_row = SubmissionRow(
            submission_id=submission.id,
            input_row=2,
            input_product_name="تخم مرغ شانه ۳۰ عددی",
            input_price=None,
            final_price=None,
        )
        db.add_all([ready_row, review_row])
        db.commit()
        db.refresh(ready_row)
        db.refresh(review_row)
        ready_matches = [
            TorobMatch(
                row_id=ready_row.id,
                source="torob",
                rank=index,
                base_prk=f"oil-{index}",
                name=f"روغن پیشنهادی {index}",
                price=330000 + index,
                image_url=None,
                product_url=None,
                is_already_added=False,
            )
            for index in range(6)
        ]
        review_match = TorobMatch(
            row_id=review_row.id,
            source="torob",
            rank=0,
            base_prk="egg-1",
            name="تخم مرغ شانه ۳۰ عددی",
            price=400000,
            image_url=None,
            product_url=None,
            is_already_added=False,
        )
        db.add_all(ready_matches + [review_match])
        db.commit()
        db.refresh(ready_matches[0])
        ready_row.selected_match_id = ready_matches[0].id
        db.add(SubmissionSelection(row_id=ready_row.id, match_id=ready_matches[0].id, final_price="330000"))
        db.commit()
        submission_id = submission.id
        ready_row_id = ready_row.id
        ready_match_id = ready_matches[0].id

    match_page = client.get(f"/submissions/{submission_id}/match")
    assert match_page.status_code == 200
    assert "1 محصول آماده تایید است و 1 محصول نیاز به بررسی دارد" in match_page.text
    assert "نیازمند بررسی" in match_page.text
    assert "آماده تایید" in match_page.text
    assert "مچ" not in match_page.text
    assert "محصول 1 کانال" in match_page.text
    assert "محصول کانال 1" not in match_page.text
    assert "قیمت از کانال تشخیص داده نشد" in match_page.text
    assert "reveal-more-btn" in match_page.text
    assert match_page.text.count("hidden data-extra-match") == 2
    assert "می‌تونی همین‌هایی که آماده تایید هستند رو اضافه کنی و بقیه رو بذاری برای کمی بعد انجام بدی." in match_page.text
    assert "data-global-price-unit" not in match_page.text
    assert "ذخیره همین‌ها و ادامه بعدا" in match_page.text

    confirm = client.post(
        f"/submissions/{submission_id}/confirm",
        data={
            f"selected_{ready_row_id}": str(ready_match_id),
            f"price_{ready_row_id}": "330000",
            "price_unit": "toman",
            "finish_mode": "continue",
        },
    )
    assert confirm.status_code == 200
    assert "1 کالا ثبت شد" in confirm.text

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission.status == "ready"
        assert submission.rows[0].submitted_at is not None
        assert submission.rows[1].submitted_at is None
        assert len(submission.batches) == 1

    remaining = client.get(f"/submissions/{submission_id}/match")
    assert "روغن جامد لادن" not in remaining.text
    assert "تخم مرغ شانه ۳۰ عددی" in remaining.text

    eitaa_page = client.get("/eitaa")
    assert "بررسی محصولات فروشگاه ایتا ناتمام مانده" in eitaa_page.text
    assert "1 محصول نیازمند بررسی مانده" in eitaa_page.text


def test_price_update_preview_groups_ready_review_new_and_preserves_batch_operation(tmp_path) -> None:
    """Human meaning: update previews must show ready updates first, review items second, new products third; submitted update rows must stay marked as price_update in the batch."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    with TestingSessionLocal() as db:
        submission = Submission(
            store_name="فروشگاه آپدیت",
            source="eitaa_update",
            source_ref="@regaal",
            operation="price_update",
            status="ready",
            price_unit="toman",
            total_rows=3,
            selected_rows=1,
        )
        db.add(submission)
        db.commit()
        db.refresh(submission)

        ready_row = SubmissionRow(
            submission_id=submission.id,
            input_row=1,
            input_product_name="کفش تست",
            input_price="750000",
            final_price="750000",
            operation="price_update",
        )
        review_row = SubmissionRow(
            submission_id=submission.id,
            input_row=2,
            input_product_name="محصول نامشخص",
            input_price="880000",
            final_price="880000",
            operation="needs_review",
            error_message="این محصول نیازمند بررسی است.",
        )
        new_row = SubmissionRow(
            submission_id=submission.id,
            input_row=3,
            input_product_name="کیف جدید",
            input_price="990000",
            final_price="990000",
            operation="add",
        )
        db.add_all([ready_row, review_row, new_row])
        db.commit()
        db.refresh(ready_row)
        db.refresh(new_row)

        ready_match = TorobMatch(
            row_id=ready_row.id,
            source="torob",
            rank=0,
            base_prk="shoe-base-rk",
            name="کفش تست ترب",
            price=650000,
        )
        new_match = TorobMatch(
            row_id=new_row.id,
            source="torob",
            rank=0,
            base_prk="bag-base-rk",
            name="کیف جدید ترب",
            price=990000,
        )
        db.add_all([ready_match, new_match])
        db.commit()
        db.refresh(ready_match)
        ready_row.selected_match_id = ready_match.id
        db.add(SubmissionSelection(row_id=ready_row.id, match_id=ready_match.id, final_price="750000"))
        db.commit()

        submission_id = submission.id
        ready_row_id = ready_row.id
        ready_match_id = ready_match.id

    match_page = client.get(f"/submissions/{submission_id}/match")

    assert match_page.status_code == 200
    assert match_page.text.index("آماده آپدیت") < match_page.text.index("نیازمند بررسی") < match_page.text.index("محصول جدید")
    assert "کفش تست ترب" in match_page.text
    assert "محصول نامشخص" in match_page.text
    assert "کیف جدید ترب" in match_page.text

    confirm = client.post(
        f"/submissions/{submission_id}/confirm",
        data={
            f"selected_{ready_row_id}": str(ready_match_id),
            f"price_{ready_row_id}": "750000",
            "price_unit": "toman",
            "finish_mode": "continue",
        },
    )

    assert confirm.status_code == 200
    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert len(submission.batches) == 1
        assert submission.batches[0].items[0].operation == "price_update"


def test_excel_price_update_uses_store_history_and_torob_for_new_products(monkeypatch, tmp_path) -> None:
    """Human meaning: an Excel update should update known store products, and turn unknown-but-searchable rows into new products."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["product_name", "price"])
    sheet.append(["کفش تست", "750000"])
    sheet.append(["کیف جدید", "990000"])
    output = BytesIO()
    workbook.save(output)

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    with TestingSessionLocal() as db:
        store = Store(name="فروشگاه آپدیت", seller_phone="09120000000")
        old_submission = Submission(
            store=store,
            store_name="فروشگاه آپدیت",
            seller_phone="09120000000",
            source="excel",
            operation="add",
            status="submitted",
            price_unit="toman",
        )
        old_row = SubmissionRow(
            submission=old_submission,
            input_row=1,
            input_product_name="کفش تست",
            input_price="650000",
            final_price="650000",
            operation="add",
        )
        old_match = TorobMatch(
            row=old_row,
            source="torob",
            rank=0,
            base_prk="shoe-base-rk",
            name="کفش تست ترب",
            price=650000,
        )
        db.add_all([store, old_submission, old_row, old_match])
        db.flush()
        old_row.selected_match_id = old_match.id
        db.add(SubmissionSelection(row_id=old_row.id, match_id=old_match.id, final_price="650000"))
        db.commit()

    monkeypatch.setattr("app.routes.seller.ProductSearchClient", FakeProductSearchClient)
    monkeypatch.setattr("app.routes.seller.SessionLocal", TestingSessionLocal)
    monkeypatch.setattr("app.routes.seller.settings.upload_dir", str(tmp_path / "uploads"))

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    response = client.post(
        "/uploads",
        data={"store_name": "فروشگاه آپدیت", "seller_phone": "09120000000", "operation": "price_update"},
        files={"file": ("update.xlsx", output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    with TestingSessionLocal() as db:
        preview = db.query(Submission).filter(Submission.operation == "price_update").one()
        assert preview.source == "excel_update"
        assert preview.status == "ready"
        assert preview.price_unit == "toman"
        assert preview.total_rows == 2
        assert preview.selected_rows == 1
        first, second = preview.rows
        assert first.operation == "price_update"
        assert first.final_price == "750000"
        assert first.selections[0].match.base_prk == "shoe-base-rk"
        assert second.operation == "add"
        assert second.final_price == "990000"
        assert second.selections == []
        assert second.matches
        preview_id = preview.id

    match_page = client.get(f"/submissions/{preview_id}/match")
    assert match_page.status_code == 200
    assert match_page.text.index("آماده آپدیت") < match_page.text.index("محصول جدید")


def test_excel_price_update_ignores_rows_when_price_did_not_change(monkeypatch, tmp_path) -> None:
    """Human meaning: uploading the same Excel again as a price update should not create fake update work."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["product_name", "price"])
    sheet.append(["test tomato paste 800g", "6500000"])
    output = BytesIO()
    workbook.save(output)

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    with TestingSessionLocal() as db:
        store = Store(name="Update Store", seller_phone="09120000000")
        old_submission = Submission(
            store=store,
            store_name="Update Store",
            seller_phone="09120000000",
            source="excel",
            operation="add",
            status="submitted",
            price_unit="rial",
        )
        old_row = SubmissionRow(
            submission=old_submission,
            input_row=1,
            input_product_name="test tomato paste 800g",
            input_price="6500000",
            final_price="650000",
            operation="add",
            submitted_at=datetime.now(timezone.utc),
        )
        old_match = TorobMatch(
            row=old_row,
            source="torob",
            rank=0,
            base_prk="tomato-paste-base-rk",
            name="test tomato paste 800g torob",
            price=650000,
        )
        db.add_all([store, old_submission, old_row, old_match])
        db.flush()
        old_row.selected_match_id = old_match.id
        db.add(SubmissionSelection(row_id=old_row.id, match_id=old_match.id, final_price="650000"))
        db.commit()

    monkeypatch.setattr("app.routes.seller.ProductSearchClient", FakeProductSearchClient)
    monkeypatch.setattr("app.routes.seller.SessionLocal", TestingSessionLocal)
    monkeypatch.setattr("app.routes.seller.settings.upload_dir", str(tmp_path / "uploads"))

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    response = client.post(
        "/uploads",
        data={"store_name": "Update Store", "seller_phone": "09120000000", "operation": "price_update"},
        files={"file": ("same-prices.xlsx", output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    with TestingSessionLocal() as db:
        preview = db.query(Submission).filter(Submission.operation == "price_update").one()
        assert preview.status == "ready"
        assert preview.selected_rows == 0
        assert preview.price_unit == "rial"
        assert preview.rows[0].operation == "unchanged"
        assert preview.rows[0].selections == []

    match_page = client.get(f"/submissions/{preview.id}/match")
    assert match_page.status_code == 200
    assert "test tomato paste 800g" not in match_page.text


def test_excel_price_update_match_page_does_not_require_price_unit_choice(monkeypatch, tmp_path) -> None:
    """Human meaning: Excel price updates are stored in Toman, so the seller must not see the Rial/Toman blocker."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["product_name", "price"])
    sheet.append(["test tomato paste 800g", "7500000"])
    output = BytesIO()
    workbook.save(output)

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    with TestingSessionLocal() as db:
        store = Store(name="Update Store", seller_phone="09120000000")
        old_submission = Submission(
            store=store,
            store_name="Update Store",
            seller_phone="09120000000",
            source="excel",
            operation="add",
            status="submitted",
            price_unit="rial",
        )
        old_row = SubmissionRow(
            submission=old_submission,
            input_row=1,
            input_product_name="test tomato paste 800g",
            input_price="6500000",
            final_price="650000",
            operation="add",
            submitted_at=datetime.now(timezone.utc),
        )
        old_match = TorobMatch(
            row=old_row,
            source="torob",
            rank=0,
            base_prk="tomato-paste-base-rk",
            name="test tomato paste 800g torob",
            price=650000,
        )
        db.add_all([store, old_submission, old_row, old_match])
        db.flush()
        old_row.selected_match_id = old_match.id
        db.add(SubmissionSelection(row_id=old_row.id, match_id=old_match.id, final_price="650000"))
        db.commit()

    monkeypatch.setattr("app.routes.seller.ProductSearchClient", FakeProductSearchClient)
    monkeypatch.setattr("app.routes.seller.SessionLocal", TestingSessionLocal)
    monkeypatch.setattr("app.routes.seller.settings.upload_dir", str(tmp_path / "uploads"))

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    response = client.post(
        "/uploads",
        data={"store_name": "Update Store", "seller_phone": "09120000000", "operation": "price_update"},
        files={"file": ("changed-price.xlsx", output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    with TestingSessionLocal() as db:
        preview = db.query(Submission).filter(Submission.operation == "price_update").one()
        preview_id = preview.id

    match_page = client.get(f"/submissions/{preview_id}/match")

    assert match_page.status_code == 200
    assert 'name="price_unit" value="rial"' in match_page.text
    assert '<div class="price-unit-card" data-price-unit-card>' not in match_page.text
    assert "قیمت‌ها ریال هست یا تومان؟" not in match_page.text
    assert "اول مشخص کن قیمت‌ها به تومان هستند یا ریال." not in match_page.text


def test_seller_can_create_eitaa_price_update_preview(monkeypatch, tmp_path) -> None:
    """Human meaning: the seller can ask Torobjan to read new Eitaa updates and receive a review page before anything is sent to Torob."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    with TestingSessionLocal() as db:
        store = Store(
            name="فروشگاه ایتا",
            seller_phone="09120000000",
            eitaa_channel_id="@regaal",
            eitaa_last_update_id=84,
        )
        old_submission = Submission(
            store=store,
            store_name="فروشگاه ایتا",
            seller_phone="09120000000",
            source="eitaa",
            source_ref="@regaal",
            operation="add",
            status="submitted",
            price_unit="toman",
        )
        old_row = SubmissionRow(
            submission=old_submission,
            input_row=1,
            input_product_name="کفش تست",
            input_price="650000",
            final_price="650000",
            source_message_id="124",
            operation="add",
        )
        old_match = TorobMatch(
            row=old_row,
            source="torob",
            rank=0,
            base_prk="shoe-base-rk",
            name="کفش تست ترب",
            price=650000,
        )
        db.add_all([store, old_submission, old_row, old_match])
        db.flush()
        old_row.selected_match_id = old_match.id
        db.add(SubmissionSelection(row_id=old_row.id, match_id=old_match.id, final_price="650000"))
        db.commit()

    class UpdateUniomClient:
        async def get_updates(self, offset=None, timeout_seconds=30):
            return [
                {
                    "update_id": 85,
                    "edited_channel_post": {
                        "message_id": 124,
                        "text": "کفش تست - 750000 تومان",
                    },
                },
                {
                    "update_id": 86,
                    "channel_post": {
                        "message_id": 125,
                        "text": "کیف جدید - 990000 تومان",
                    },
                },
            ]

        async def close(self):
            return None

    monkeypatch.setattr("app.routes.seller.UniomClient", UpdateUniomClient)
    monkeypatch.setattr("app.routes.seller.ProductSearchClient", FakeProductSearchClient)
    monkeypatch.setattr("app.routes.seller.settings.uniom_bot_token", "test-token")

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    response = client.post(
        "/eitaa/update",
        data={"store_name": "فروشگاه ایتا", "seller_phone": "09120000000", "channel_id": "regaal"},
    )

    assert response.status_code == 200
    assert "آماده آپدیت" in response.text
    assert "محصول جدید" in response.text
    with TestingSessionLocal() as db:
        preview = db.query(Submission).filter(Submission.operation == "price_update").one()
        assert preview.source == "eitaa_update"
        assert preview.store.eitaa_last_update_id == 86
        assert [row.operation for row in preview.rows] == ["price_update", "add"]
        assert preview.rows[0].selections[0].match.base_prk == "shoe-base-rk"
        assert preview.rows[1].matches


def test_admin_does_not_send_price_update_batches_with_add_endpoint(monkeypatch, tmp_path) -> None:
    """Human meaning: until Torob gives us an upsert endpoint, update-price batches must not be sent through the add-only endpoint."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    with TestingSessionLocal() as db:
        submission = Submission(
            store_name="فروشگاه آپدیت",
            shop_id="411488",
            source="excel_update",
            operation="price_update",
            status="submitted",
            price_unit="toman",
        )
        row = SubmissionRow(
            submission=submission,
            input_row=1,
            input_product_name="کفش تست",
            input_price="750000",
            final_price="750000",
            operation="price_update",
            submitted_at=datetime.now(timezone.utc),
        )
        match = TorobMatch(row=row, source="torob", rank=0, base_prk="shoe-base-rk", name="کفش تست ترب", price=650000)
        batch = SubmissionBatch(submission=submission, status="pending")
        item = SubmissionBatchItem(batch=batch, row=row, match=match, operation="price_update", final_price="750000")
        db.add_all([submission, row, match, batch, item])
        db.commit()
        submission_id = submission.id
        batch_id = batch.id

    monkeypatch.setattr("app.routes.admin.settings.admin_password", "secret")
    monkeypatch.setattr("app.routes.admin.settings.session_secret", "test-cookie")
    monkeypatch.setattr("app.routes.admin.settings.torob_bulk_add_key", "bulk-test-key")
    monkeypatch.setattr("app.routes.admin.TorobBulkAddClient", FakeTorobBulkAddClient)
    FakeTorobBulkAddClient.calls = []

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)
    assert client.post("/admin/login", data={"password": "secret"}, follow_redirects=False).status_code == 303

    response = client.post(f"/admin/submissions/{submission_id}/batches/{batch_id}/send-torob", follow_redirects=False)

    assert response.status_code == 303
    assert FakeTorobBulkAddClient.calls == []
    with TestingSessionLocal() as db:
        batch = db.query(SubmissionBatch).first()
        assert batch.status == "failed"
        assert "upsert" in batch.torob_error_message


def test_old_eitaa_draft_forces_toman_without_showing_unit_controls(tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    with TestingSessionLocal() as db:
        submission = Submission(
            store_name="فروشگاه قدیمی ایتا",
            source="eitaa",
            source_ref="@legacy",
            status="ready",
            price_unit=None,
            total_rows=1,
            selected_rows=1,
        )
        db.add(submission)
        db.commit()
        db.refresh(submission)
        row = SubmissionRow(
            submission_id=submission.id,
            input_row=1,
            input_product_name="روغن جامد لادن",
            input_price="330000",
            final_price="330000",
        )
        db.add(row)
        db.commit()
        db.refresh(row)
        match = TorobMatch(
            row_id=row.id,
            source="torob",
            rank=0,
            base_prk="legacy-oil",
            name="روغن جامد لادن",
            price=330000,
            image_url=None,
            product_url=None,
            is_already_added=False,
        )
        db.add(match)
        db.commit()
        db.refresh(match)
        row.selected_match_id = match.id
        db.add(SubmissionSelection(row_id=row.id, match_id=match.id, final_price="330000"))
        db.commit()
        submission_id = submission.id
        row_id = row.id
        match_id = match.id

    match_page = client.get(f"/submissions/{submission_id}/match")

    assert match_page.status_code == 200
    assert 'name="price_unit" value="toman"' in match_page.text
    assert 'data-global-price-unit' not in match_page.text
    assert 'data-price-unit-error-template' not in match_page.text
    assert 'data-price-unit-option="toman"' not in match_page.text
    assert 'data-price-unit-option="rial"' not in match_page.text
    assert "اول مشخص کن قیمت‌ها به تومان هستند یا ریال." not in match_page.text

    draft = client.post(
        f"/submissions/{submission_id}/draft",
        data={
            f"selected_{row_id}": str(match_id),
            f"price_{row_id}": "1,200,000",
        },
    )
    assert draft.status_code == 200

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission.price_unit == "toman"
        assert submission.rows[0].final_price == "1200000"

    confirm = client.post(
        f"/submissions/{submission_id}/confirm",
        data={
            f"selected_{row_id}": str(match_id),
            f"price_{row_id}": "1,500,000",
            "price_unit": "rial",
        },
    )
    assert confirm.status_code == 200

    with TestingSessionLocal() as db:
        submission = db.query(Submission).first()
        assert submission.price_unit == "toman"
        assert submission.rows[0].final_price == "1500000"


def test_eitaa_resume_card_counts_review_rows_not_ready_rows(tmp_path) -> None:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    with TestingSessionLocal() as db:
        submission = Submission(
            store_name="تست تست",
            source="eitaa",
            source_ref="@timanic_shop",
            status="ready",
            total_rows=400,
        )
        db.add(submission)
        db.commit()
        db.refresh(submission)
        for index in range(1, 401):
            row = SubmissionRow(
                submission_id=submission.id,
                input_row=index,
                input_product_name=f"محصول {index}",
                input_price="100000",
                final_price="100000" if index <= 148 else None,
            )
            db.add(row)
            db.flush()
            match = TorobMatch(
                row_id=row.id,
                source="torob",
                rank=0,
                base_prk=f"base-{index}",
                name=f"محصول {index}",
                price=100000,
                image_url=None,
                product_url=None,
                is_already_added=False,
            )
            db.add(match)
            db.flush()
            if index <= 148:
                row.selected_match_id = match.id
                db.add(SubmissionSelection(row_id=row.id, match_id=match.id, final_price="100000"))
        db.commit()
        submission_id = submission.id

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)
    client.cookies.set("torobjan_latest_submission", str(submission_id))

    response = client.get("/eitaa")

    assert response.status_code == 200
    assert "252 محصول نیازمند بررسی مانده" in response.text
    assert "400 محصول" not in response.text


def test_index_has_loading_submit_state() -> None:
    app = create_app()
    client = TestClient(app)

    home = client.get("/")

    assert home.status_code == 200
    assert 'href="/excel"' in home.text
    assert 'href="/eitaa"' in home.text
    assert "upload-form" not in home.text
    assert "وارد کردن دسته جمعی محصولات به ترب، برای فروشندگان حضوری" in home.text
    assert "آیدی کانالت را بده" in home.text

    response = client.get("/excel")

    assert response.status_code == 200
    assert "upload-form" in response.text
    assert "file-drop" in response.text
    assert "هنوز فایلی انتخاب نشده است" in response.text
    assert "برای شروع، فایل اکسل محصولات را انتخاب کن" in response.text
    assert "در حال پردازش فایل" in response.text
    assert "fetch(form.action" in response.text
    assert "شماره موبایل، اختیاری" in response.text
    assert "وارد کردن محصولات با فایل اکسل" in response.text
    assert "این نسخه آزمایشی هست؛ لطفا شمارت رو بذار" in response.text
    assert "clarity.ms/tag" not in response.text


def test_eitaa_primary_submit_posts_to_import_route() -> None:
    """Human meaning: the main Eitaa button must post to /eitaa/import, not accidentally POST the /eitaa page and get 405."""
    app = create_app()
    client = TestClient(app)

    response = client.get("/eitaa")

    assert response.status_code == 200
    assert 'action="/eitaa/import"' in response.text
    assert 'formaction="/eitaa/import"' in response.text
    assert 'formaction="/eitaa/update"' in response.text
    assert 'getAttribute("formaction")' in response.text


def test_health_endpoint() -> None:
    app = create_app()
    client = TestClient(app)

    response = client.get("/health")

    assert response.status_code == 200
    assert response.text == "ok"
