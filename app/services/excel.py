from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
import re
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.models import Submission, SubmissionBatch, SubmissionRow

MAX_ROWS = 500

PRODUCT_NAME_HEADERS = {
    "product_name",
    "product name",
    "name",
    "title",
    "item",
    "goods",
    "نام محصول",
    "نام کالا",
    "کالا",
    "محصول",
    "عنوان محصول",
    "عنوان کالا",
    "شرح کالا",
    "شرح محصول",
}
PRICE_HEADERS = {"price", "قیمت", "قیمت فروش", "قیمت تومان", "قیمت نهایی", "مبلغ", "فی", "فی فروش"}
BARCODE_HEADERS = {"barcode", "bar code", "bar_code", "کد کالا", "بارکد", "کد محصول", "شناسه کالا"}
BRAND_HEADERS = {"brand", "برند", "نام برند", "مارک"}
DESCRIPTION_HEADERS = {"description", "desc", "توضیحات", "شرح", "مشخصات"}


@dataclass(frozen=True)
class ParsedExcelRow:
    input_row: int
    product_name: str | None
    price: str | None
    barcode: str | None
    brand: str | None
    description: str | None
    error_message: str | None = None


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\ufeff", "")
    text = text.replace("\u200c", " ")
    text = text.replace("ي", "ی").replace("ك", "ک")
    text = text.replace("_", " ")
    text = re.sub(r"[\s:：؛،,\-–—/\\|()\[\]]+", " ", text)
    return text.strip()


def clean_cell(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def parse_price(value: object) -> str | None:
    text = clean_cell(value)
    if not text:
        return None
    normalized = text.replace(",", "").replace("٬", "").replace("٫", "").replace(" ", "")
    return normalized if normalized.isdigit() else text


def _find_column(headers: list[str], candidates: set[str]) -> int | None:
    normalized_candidates = {normalize_header(candidate) for candidate in candidates}
    for index, header in enumerate(headers):
        if header in normalized_candidates:
            return index
    for index, header in enumerate(headers):
        if any(candidate and candidate in header for candidate in normalized_candidates):
            return index
    return None


def parse_products_excel(content: bytes, filename: str) -> list[ParsedExcelRow]:
    if filename.lower().endswith(".xls"):
        return _parse_products_rows(_read_xls_rows(content))
    return parse_products_xlsx(content)


def parse_products_xlsx(content: bytes) -> list[ParsedExcelRow]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    return _parse_products_rows(rows)


def _read_xls_rows(content: bytes) -> list[list[object]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ExcelParseError("برای خواندن فایل xls قدیمی، پکیج xlrd باید نصب باشد.") from exc
    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    return [sheet.row_values(index) for index in range(sheet.nrows)]


def _parse_products_rows(rows: list[list[object]]) -> list[ParsedExcelRow]:
    if not rows:
        return []

    header_index = _find_header_row(rows)
    if header_index is None:
        return [
            ParsedExcelRow(
                input_row=1,
                product_name=None,
                price=None,
                barcode=None,
                brand=None,
                description=None,
                error_message="ستون نام محصول پیدا نشد. ستون‌هایی مثل «نام کالا»، «نام محصول» یا «محصول» را در فایل بگذار.",
            )
        ]

    headers = [normalize_header(value) for value in rows[header_index]]
    name_col = _find_column(headers, PRODUCT_NAME_HEADERS)
    price_col = _find_column(headers, PRICE_HEADERS)
    barcode_col = _find_column(headers, BARCODE_HEADERS)
    brand_col = _find_column(headers, BRAND_HEADERS)
    desc_col = _find_column(headers, DESCRIPTION_HEADERS)

    parsed: list[ParsedExcelRow] = []
    if name_col is None:
        return [
            ParsedExcelRow(
                input_row=header_index + 1,
                product_name=None,
                price=None,
                barcode=None,
                brand=None,
                description=None,
                error_message="ستون نام محصول پیدا نشد. ستون‌هایی مثل «نام کالا»، «نام محصول» یا «محصول» را در فایل بگذار.",
            )
        ]

    data_rows = rows[header_index + 1 : header_index + 1 + MAX_ROWS]
    for excel_index, row in enumerate(data_rows, start=header_index + 2):
        product_name = clean_cell(row[name_col] if name_col < len(row) else None)
        price = parse_price(row[price_col] if price_col is not None and price_col < len(row) else None)
        barcode = clean_cell(row[barcode_col] if barcode_col is not None and barcode_col < len(row) else None)
        brand = clean_cell(row[brand_col] if brand_col is not None and brand_col < len(row) else None)
        description = clean_cell(row[desc_col] if desc_col is not None and desc_col < len(row) else None)
        if not any([product_name, price, barcode, brand, description]):
            continue
        parsed.append(
            ParsedExcelRow(
                input_row=excel_index,
                product_name=product_name,
                price=price,
                barcode=barcode,
                brand=brand,
                description=description,
                error_message=None if product_name else "نام محصول خالی است.",
            )
        )

    if len(rows) - header_index - 1 > MAX_ROWS:
        parsed.append(
            ParsedExcelRow(
                input_row=MAX_ROWS + 2,
                product_name=None,
                price=None,
                barcode=None,
                brand=None,
                description=None,
                error_message=f"فقط {MAX_ROWS} ردیف اول پردازش شد.",
            )
        )
    return parsed


def _find_header_row(rows: list[list[object]]) -> int | None:
    best_index: int | None = None
    best_score = 0
    for index, row in enumerate(rows[:10]):
        headers = [normalize_header(value) for value in row]
        score = 0
        for candidates in [PRODUCT_NAME_HEADERS, PRICE_HEADERS, BARCODE_HEADERS, BRAND_HEADERS, DESCRIPTION_HEADERS]:
            if _find_column(headers, candidates) is not None:
                score += 1
        if score > best_score:
            best_index = index
            best_score = score
    return best_index if best_score > 0 else None


def build_template_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "products"
    headers = ["product_name", "price", "barcode", "brand", "description"]
    sheet.append(headers)
    sheet.append(["رب گوجه فرنگی روژین ۸۰۰ گرم", "165000", "6260000000000", "روژین", ""])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="D73948")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_export_xlsx(submission: Submission, rows: Iterable[SubmissionRow]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "torob_export"
    _append_export_headers(sheet)

    created_at = _format_datetime(submission.created_at)
    for row in rows:
        for selection in row.selections:
            if _is_exportable_selection(selection.match, selection.final_price):
                _append_export_row(sheet, submission, row, selection.match, selection.final_price, created_at)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_batch_export_xlsx(submission: Submission, batch: SubmissionBatch) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "torob_export"
    _append_export_headers(sheet)

    created_at = _format_datetime(batch.created_at)
    for item in batch.items:
        if _is_exportable_selection(item.match, item.final_price):
            _append_export_row(sheet, submission, item.row, item.match, item.final_price, created_at)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _append_export_headers(sheet) -> None:
    headers = [
        "submission_id",
        "store_name",
        "seller_phone",
        "shop_id",
        "input_row",
        "input_product_name",
        "input_price",
        "selected_source",
        "selected_base_prk",
        "selected_torob_name",
        "selected_torob_price",
        "final_price",
        "image_url",
        "torob_product_url",
        "created_at",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="D73948")


def _is_exportable_selection(match, final_price: str | None) -> bool:
    if match is None or not final_price:
        return False
    normalized = parse_price(final_price)
    if not normalized:
        return False
    digits = re.sub(r"\D+", "", normalized)
    return bool(digits) and int(digits) > 0


def _append_export_row(sheet, submission: Submission, row: SubmissionRow, match, final_price: str | None, created_at: str | None) -> None:
    sheet.append(
        [
            submission.id,
            submission.store_name,
            submission.seller_phone,
            submission.shop_id,
            row.input_row,
            row.input_product_name,
            row.input_price,
            (match.source or "torob") if match else None,
            match.base_prk if match else None,
            match.name if match else None,
            match.price if match else None,
            final_price,
            match.image_url if match else None,
            match.product_url if match else None,
            created_at,
        ]
    )


def _format_datetime(value: datetime | None) -> str | None:
    return value.isoformat() if value else None


class ExcelParseError(RuntimeError):
    pass
